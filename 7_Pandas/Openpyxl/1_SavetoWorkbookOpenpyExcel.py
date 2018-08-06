# -*- coding: utf-8 -*-
"""
Created on Fri Jun  1 12:48:15 2018

@author: SilverDoe
"""

import openpyxl

filename = 'E:\\trashcode\\dem.xlsx'
wb = openpyxl.load_workbook(filename=filename)
ws = wb.worksheets[0]
ws['A1'] = 1
ws['B1'] = 22
ws['C1'] = 333
ws['D1'] = 4444
ws['E1'] = 55555
ws['F1'] = 666666
wb.save(filename)
